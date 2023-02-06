from flask import jsonify
import openpyxl, unicodedata
from .dbmodels import *
from .dbschemas import *
from .dbsettings import new_Scoped_session
import pandas as pd, numpy as np
from .config import STORAGE_PATH
import requests, logging
INITIALDATA = "components/initial_data/"

def InitialDataFile(name: str):
    return INITIALDATA + name + ".xlsx"

def TruncateTables(tables: list):
    Session = new_Scoped_session()
    Session.execute("SET FOREIGN_KEY_CHECKS=0;")
    Session.commit()
    for table in tables:
        try:
            Session.execute("TRUNCATE TABLE {};".format(table))
        except Exception as e:
            if "1146" in str(e):
                print("Table {} does not exist".format(table))
            else:
                print("Truncate table {} error: {}".format(table,e))

    Session.execute("SET FOREIGN_KEY_CHECKS=1;")
    Session.commit()
    Session.close()


def remove_accents(input_str = None):
    print(f"Debug: {input_str}")
    if input_str == None:
        return ''
    nkfd_form = unicodedata.normalize('NFKD', input_str)
    only_ascii = nkfd_form.encode('ASCII', 'ignore')
    return only_ascii.decode()

def InsertLocation1():
    
    TruncateTables({"CITY", "DISTRICT", "WARD"})
    wb = openpyxl.load_workbook(InitialDataFile("Locations"))

    sheet = wb.active
    i = 1
    c = 0
    d = 0
    w = 0
    Session = new_Scoped_session()
    cities = []
    districts = []
    wards = []
    
    while True:
        i += 1
        cityName = sheet.cell(row=i, column=1).value
        if cityName == None:
            break
        districtName = sheet.cell(row=i, column=3).value
        wardName = sheet.cell(row=i, column=5).value

        if cityName not in cities:
            city = City(Name=cityName)
            Session.add(city)
            cities.append(cityName)
            c += 1
        
        cityID = c

        tempDistrict = {
            "Name": districtName,
            "CityID": cityID
        }

        if tempDistrict not in districts:
            district = District(Name=districtName, ID_City=cityID)
            Session.add(district)
            districts.append(tempDistrict)
            d += 1

        districtID = d

        if wardName == None:
            continue

        tempWard = {
            "Name": wardName,
            "DistrictID": districtID
        }

        if tempWard not in wards:
            ward = Ward(Name=wardName, ID_District=districtID)
            Session.add(ward)
            wards.append(tempWard)
            w += 1

    Session.commit()
    Session.close()
    return "Inserted {} cities, {} districts, {} wards".format(c, d, w)


def InsertLocation2():
    TruncateTables({"CITY", "DISTRICT", "WARD"})
    city_table = pd.read_csv(INITIALDATA + 'Locations_City.csv', index_col=False)
    district_table = pd.read_csv(INITIALDATA + 'Locations_District.csv', index_col=False)
    ward_table = pd.read_csv(INITIALDATA + 'Locations_Ward.csv', index_col=False)
    Session = new_Scoped_session()
    try:
        for table in [city_table, district_table, ward_table]:
            for i, row in table.iterrows():
                if table is city_table: new_row = City(ID=row['City_UID'], Name=row['City'])
                elif table is district_table: new_row = District(ID=row['District_UID'], Name=row['District'], ID_City=row['City_UID'])
                else: new_row = Ward(ID=row['Ward_UID'], Name=row['Ward'], ID_District=row['District_UID'])
                Session.add(new_row)
            Session.flush()
        Session.commit()
        return "Inserted {} cities, {} districts, {} wards".format(city_table.shape[0], district_table.shape[0], ward_table.shape[0])
    except Exception as e:
        Session.rollback()
        return "Error: " + str(e)
    
    
def InsertPermission():
    TruncateTables({"Permission"})
    wb = openpyxl.load_workbook(InitialDataFile("Permissions"))

    sheet = wb.active
    i = 1
    p = 0

    Session = new_Scoped_session()
    while True:
        i += 1
        name = sheet.cell(row=i, column=1).value
        if name == None:
            break
        perm = Session.query(Permission).filter(Permission.Name == name).first()
        if perm == None or perm.Name != name:
            perm = Permission(Name=name)
            Session.add(perm)
            p += 1

    Session.commit()
    Session.close()

    return "Inserted {} permissions".format(p)

def InsertImageType():
    TruncateTables({"ImageType"})
    wb = openpyxl.load_workbook(InitialDataFile("ImageTypes"))
    Session = new_Scoped_session()
    sheet = wb.active
    i = 1
    it = 0
    while True:
        i += 1
        name = sheet.cell(row=i, column=1).value
        if name == None:
            break
        imgType = Session.query(ImageType).filter(ImageType.Name == name).first()
        if imgType == None or imgType.Name != name:
            imgType = ImageType(Name=name)
            Session.add(imgType)
            it += 1

    Session.commit()
    Session.close()

    return "Inserted {} image types".format(it)


def GetImageFolder(image_type):
    if image_type == 1: folder = 'post/'
    elif image_type == 2: folder = 'logo/'
    elif image_type == 3: folder = 'user/'
    elif image_type == 4: folder = 'identity/'
    return folder


def SaveImageFromURL(Session, url, image_type):
    r = requests.get(url)
    if r.status_code==200:
        filename = url.split('/')[-1]
        image = Image(Filename=filename, ID_ImageType=image_type)
        Session.add(image)
        Session.flush()
        ext = image.Filename.split('.')[-1]
        image.Filename = str(image.ID) + '.' + ext
        open(STORAGE_PATH + GetImageFolder(image_type) + image.Filename, "wb").write(r.content)
        Session.flush()
        return [True, image.ID]
    return [False, "Invalid response"]


def SaveImage(Session, file, image_type):
    if file.filename == "":
        return [False, "No file selected"]
    ext = file.filename.split('.')[-1]
    if ext not in ['jpg', 'jpeg', 'png']:
        return [False, "File extension not supported"]
    try:
        new_image = dbm.Image(Filename = "", ID_ImageType=image_type)
        Session.add(new_image)
        Session.flush()
        new_image.Filename = str(new_image.ID) + '.' + ext
        Session.commit()
        file.save(os.path.join(STORAGE_PATH, GetImageFolder(image_type), str(new_image.ID) + '.' + ext))
        return [True, new_image]
    except Exception as e:
        Session.rollback()
        return [False, str(e)]
  

def InsertVehicleSupportTable():
    TruncateTables({"VEHICLEBRAND", "VEHICLELINEUP", "VEHICLETYPE", "VEHICLECONDITION", "COLOR"})
    vehiclebrand_table = pd.read_csv(INITIALDATA + 'VehicleBrand.csv', index_col=False)
    vehiclelineup_table = pd.read_csv(INITIALDATA + 'VehicleLineup.csv', index_col=False)
    vehicletype_table = pd.read_csv(INITIALDATA + 'VehicleType.csv', index_col=False)
    color_table = pd.read_csv(INITIALDATA + 'Color.csv', index_col=False)
    vehiclecondition_table = pd.read_csv(INITIALDATA + 'VehicleCondition.csv', index_col=False)
    
    Session = new_Scoped_session()
    try:
        # logolist = Session.query(Image).filter(Image.ID_ImageType==2).all()
        # for item in logolist:
        #     Session.delete(item)
        # Session.flush()
        
        for table in [vehiclebrand_table, vehicletype_table, color_table, vehiclecondition_table, vehiclelineup_table]:
            print("Table switched")
            for i, row in table.iterrows():
                if table is vehiclebrand_table: 
                    new_row = VehicleBrand(ID=row['id'], Name=row['name'])
                    # image_url = row['logo']
                    # output = SaveImageFromURL(Session=Session, url=image_url, image_type=2)
                    # if output[0]:
                    #     new_row = VehicleBrand(ID=row['id'], Name=row['name'], ID_Image=output[1])
                    # else: 
                    #     new_row = VehicleBrand(ID=row['id'], Name=row['name'])
                elif table is vehicletype_table: 
                    new_row = VehicleType(ID=row['id'], Type=row['name'])
                elif table is color_table:
                    new_row = Color(ID=row['id'], Name=row['name'], Color_hex=row['color_code'])
                elif table is vehiclecondition_table:
                    new_row = VehicleCondition(ID=row['id'], Condition=row['name'])
                else: 
                    new_row = VehicleLineup(ID=row['id'], Lineup=row['name'], ID_VehicleBrand=row['brand_id'])
                Session.add(new_row)
            Session.flush()
        Session.commit()
        return "Insert completed"
    except Exception as e:
        Session.rollback()
        return "Error: " + str(e)
    
    
def SetupAccount(Session, a_email, a_username, a_password, a_type, a_permission, i_name, i_phone=None, i_image=None):
   try:
      account = dbm.Account(Username=a_username, Password=a_password, Email=a_email, Account_type=a_type, ID_Permission=a_permission)
      accountinfo = dbm.AccountInfo(Name=i_name, Phone_number=i_phone)
      accountstat = dbm.AccountStat()
      if i_image != None: 
         output = SaveImageFromURL(Session, i_image, 3)
         if output[0]: accountinfo.ID_Image_Profile = output[1]
      Session.add(accountstat)
      Session.add(accountinfo)
      Session.flush()
      account.ID_AccountStat = accountstat.ID
      account.ID_AccountInfo = accountinfo.ID
      Session.add(account)
      Session.flush()
      return [True, account]
   except Exception as e:
      return [False, str(e)]

    
def InsertTestAccount(Session):
    try:
        oldtestuser = Session.query(Account).filter(Account.Account_type==3).scalar()
        if oldtestuser is not None: 
            Session.delete(oldtestuser)
            Session.flush()
        output = SetupAccount(Session, "testuser@email.com", "testuser", "testuserpassword", 3, 4, "Mobike Testuser", "0123456789", None)
        if output[0]:
            new_address = dbm.Address(
                Detail_address="Test address", 
                ID_AccountInfo=output[1].ID_AccountInfo, 
                ID_City=1, 
                ID_District=1, 
                ID_Ward=1)
            Session.add(new_address)
            Session.flush()
            return [True, output[1], new_address]
        else:
            return [False, output[1]]
    except Exception as e:
        return [False, str(e)]
    
    
def InsertTestdata():
    # TruncateTables({"VEHICLEINFO", "POST"})
    posts = pd.read_csv(INITIALDATA + 'PostCompleted.csv', index_col=False)
    posts = posts.replace({np.nan: None})
    Session = new_Scoped_session()
    try:
        output = InsertTestAccount(Session)
        if output[0]:
            Session.commit()
            logging.warning(f"Account created, id = {output[1].ID}")
            logging.warning(f"Address created, id = {output[2].ID}")
            acc = output[1]
            address = output[2]
            
            for index, row in posts.iterrows():
                temp_Session = new_Scoped_session()
                try:
                    logging.warning(f"Adding {index}")
                    new_vehicleinfo = VehicleInfo(
                        Vehicle_name = row['Vehicle_name'],
                        Odometer = row['Odometer'],
                        License_plate = row['License_plate'],
                        Manufacture_year = row['Manufacture_year'],
                        Cubic_power = row['Cubic_power'],
                        ID_VehicleBrand = row['ID_VehicleBrand'],
                        ID_VehicleLineup = row['ID_VehicleLineup'],
                        ID_VehicleType = row['ID_VehicleType'],
                        ID_Condition = row['Condition'],
                        ID_Color = row['ID_Color']
                    )
                    new_poststat = PostStat()
                    temp_Session.add(new_vehicleinfo)
                    temp_Session.add(new_poststat)
                    temp_Session.flush()
                    # add placeholder image to post
                    new_post = Post(
                        Title = row['Title'],
                        Content = "Test content",
                        Pricetag = row['Pricetag'],
                        ID_Account = acc.ID,
                        ID_Address = address.ID,
                        ID_PostStat = new_poststat.ID,
                        ID_VehicleInfo = new_vehicleinfo.ID
                    )
                    temp_Session.add(new_post)
                    temp_Session.flush()
                    print(new_post.ID)
                    new_poststatus = PostStatus(
                        Status = 1,
                        ID_Post = new_post.ID,
                        Information = "Added post for testing."
                    )
                    temp_Session.add(new_poststatus)
                    temp_Session.flush()
                    print(new_poststatus.ID)
                    temp_Session.commit()
                except Exception as e:
                    temp_Session.rollback()
                    logging.warning(f"Post index {index} failed, skipped. Error: {str(e)}")
        else: 
            Session.rollback()
            return [False, output[1]]
        return [True, acc.ID]
    except Exception as e:
        return [False, str(e)]